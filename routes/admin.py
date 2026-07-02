"""Admin blueprint — onboards agencies, projects, skills, rate cards, workers."""
import os, re, tempfile
from collections import Counter
from datetime import date
from decimal import Decimal

from flask import Blueprint, render_template, request, redirect, url_for, flash, abort, jsonify
from flask_login import login_required, current_user

from models import (
    db, User, Agency, Project, Skill, RateCard, Worker,
    Attendance, AttendanceApproval, Invoice,
    ROLE_ADMIN, ROLE_PROCAM_REP, ROLE_VENDOR_REP, ROLE_WORKER,
    APPROVER_DUAL, APPROVER_NONE, CAT_SKILLED,
    Office, SelfAttendanceSettings, ATT_TYPE_SELF, SIDE_MANAGER, SIDE_HR,
)
from utils import role_required, parse_date, to_ist, worker_can_self_attend

bp = Blueprint("admin", __name__)


# ---------------------------------------------------------------------------
# NPR (Non-Payroll) worker creation
# ---------------------------------------------------------------------------
# Worker code format requested by HR:
#   NPR  +  4-digit sequence  +  4-digit year
# Example: NPR00012026, NPR00022026, ...  Sequence restarts each calendar year.
NPR_CODE_RE = re.compile(r"^NPR(\d{4})(\d{4})$")
NPR_AGENCY_CODE = "NPR"
NPR_AGENCY_NAME = "Non-Payroll Workers"


def _next_npr_code(year: int) -> str:
    """Find the largest existing NPR####<year> code and return the next one."""
    prefix = "NPR"
    suffix = str(year)
    existing = (Worker.query
                .filter(Worker.code.like(f"{prefix}%{suffix}"))
                .all())
    max_seq = 0
    for w in existing:
        m = NPR_CODE_RE.match(w.code or "")
        if m and int(m.group(2)) == year:
            max_seq = max(max_seq, int(m.group(1)))
    return f"NPR{max_seq + 1:04d}{year}"


def _ensure_npr_agency() -> Agency:
    """Find/create the umbrella NPR agency (dual-approval mode)."""
    a = Agency.query.filter_by(code=NPR_AGENCY_CODE).first()
    if a:
        return a
    a = Agency(
        code=NPR_AGENCY_CODE, name=NPR_AGENCY_NAME,
        contact_person="HR Admin", email="hr@procamlogistics.com",
        address="Various vendors", gstin=None, pan=None,
        default_gst_rate=Decimal("18.00"), default_tds_rate=Decimal("2.00"),
        onboarded_on=date.today(), is_active=True,
        approver_mode=APPROVER_DUAL,
    )
    db.session.add(a); db.session.flush()
    return a


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------
@bp.route("/")
@login_required
@role_required(ROLE_ADMIN)
def dashboard():
    stats = {
        "agencies": Agency.query.filter_by(is_active=True).count(),
        "projects": Project.query.filter_by(is_active=True).count(),
        "workers": Worker.query.filter_by(is_active=True).count(),
        "invoices": Invoice.query.count(),
        "attendance_today": Attendance.query.filter_by(work_date=date.today()).count(),
    }
    return render_template("admin/dashboard.html", stats=stats)


# ---------------------------------------------------------------------------
# Agencies
# ---------------------------------------------------------------------------
@bp.route("/agencies")
@login_required
@role_required(ROLE_ADMIN)
def agencies():
    rows = Agency.query.order_by(Agency.name).all()
    return render_template("admin/agencies.html", rows=rows)


@bp.route("/agencies/new", methods=["GET", "POST"])
@login_required
@role_required(ROLE_ADMIN)
def agency_new():
    if request.method == "POST":
        f = request.form
        a = Agency(
            code=(f.get("code") or "").strip(),
            name=(f.get("name") or "").strip(),
            contact_person=f.get("contact_person") or "",
            email=f.get("email") or "",
            phone=f.get("phone") or "",
            address=f.get("address") or "",
            gstin=f.get("gstin") or "",
            pan=f.get("pan") or "",
            bank_name=f.get("bank_name") or "",
            account_no=f.get("account_no") or "",
            ifsc=f.get("ifsc") or "",
            default_tds_rate=Decimal(f.get("default_tds_rate") or "2.00"),
            default_gst_rate=Decimal(f.get("default_gst_rate") or "18.00"),
            onboarded_on=parse_date(f.get("onboarded_on")) or date.today(),
        )
        db.session.add(a)
        db.session.commit()
        flash(f"Agency '{a.name}' onboarded.", "success")
        return redirect(url_for("admin.agencies"))
    return render_template("admin/agency_form.html", a=None)


@bp.route("/agencies/<int:aid>/edit", methods=["GET", "POST"])
@login_required
@role_required(ROLE_ADMIN)
def agency_edit(aid):
    a = Agency.query.get_or_404(aid)
    if request.method == "POST":
        f = request.form
        a.code = (f.get("code") or "").strip()
        a.name = (f.get("name") or "").strip()
        a.contact_person = f.get("contact_person") or ""
        a.email = f.get("email") or ""
        a.phone = f.get("phone") or ""
        a.address = f.get("address") or ""
        a.gstin = f.get("gstin") or ""
        a.pan = f.get("pan") or ""
        a.bank_name = f.get("bank_name") or ""
        a.account_no = f.get("account_no") or ""
        a.ifsc = f.get("ifsc") or ""
        a.default_tds_rate = Decimal(f.get("default_tds_rate") or "2.00")
        a.default_gst_rate = Decimal(f.get("default_gst_rate") or "18.00")
        a.is_active = bool(f.get("is_active"))
        db.session.commit()
        flash("Agency updated.", "success")
        return redirect(url_for("admin.agencies"))
    return render_template("admin/agency_form.html", a=a)


# ---------------------------------------------------------------------------
# Projects
# ---------------------------------------------------------------------------
@bp.route("/projects")
@login_required
@role_required(ROLE_ADMIN)
def projects():
    rows = Project.query.order_by(Project.name).all()
    return render_template("admin/projects.html", rows=rows)


@bp.route("/projects/new", methods=["GET", "POST"])
@login_required
@role_required(ROLE_ADMIN)
def project_new():
    procam_reps = User.query.filter_by(role=ROLE_PROCAM_REP, is_active=True).all()
    if request.method == "POST":
        f = request.form
        p = Project(
            code=(f.get("code") or "").strip(),
            name=(f.get("name") or "").strip(),
            client_name=f.get("client_name") or "",
            location=f.get("location") or "",
            start_date=parse_date(f.get("start_date")),
            end_date=parse_date(f.get("end_date")),
            procam_rep_id=int(f.get("procam_rep_id")) if f.get("procam_rep_id") else None,
        )
        db.session.add(p)
        db.session.commit()
        flash(f"Project '{p.name}' created.", "success")
        return redirect(url_for("admin.projects"))
    return render_template("admin/project_form.html", p=None, procam_reps=procam_reps)


@bp.route("/projects/<int:pid>/edit", methods=["GET", "POST"])
@login_required
@role_required(ROLE_ADMIN)
def project_edit(pid):
    p = Project.query.get_or_404(pid)
    procam_reps = User.query.filter_by(role=ROLE_PROCAM_REP, is_active=True).all()
    if request.method == "POST":
        f = request.form
        p.code = (f.get("code") or "").strip()
        p.name = (f.get("name") or "").strip()
        p.client_name = f.get("client_name") or ""
        p.location = f.get("location") or ""
        p.start_date = parse_date(f.get("start_date"))
        p.end_date = parse_date(f.get("end_date"))
        p.procam_rep_id = int(f.get("procam_rep_id")) if f.get("procam_rep_id") else None
        p.is_active = bool(f.get("is_active"))
        db.session.commit()
        flash("Project updated.", "success")
        return redirect(url_for("admin.projects"))
    return render_template("admin/project_form.html", p=p, procam_reps=procam_reps)


# ---------------------------------------------------------------------------
# Skills
# ---------------------------------------------------------------------------
@bp.route("/skills", methods=["GET", "POST"])
@login_required
@role_required(ROLE_ADMIN)
def skills():
    from models import WORKER_CATEGORIES, CAT_UNSKILLED
    if request.method == "POST":
        # Either add a new skill or update an existing one's category
        edit_id = request.form.get("edit_id", type=int)
        if edit_id:
            sk = Skill.query.get_or_404(edit_id)
            new_cat = (request.form.get("category") or CAT_UNSKILLED).strip()
            if new_cat not in WORKER_CATEGORIES:
                flash("Invalid category.", "warning")
            else:
                sk.category = new_cat
                db.session.commit()
                flash(f"Skill '{sk.name}' → {new_cat}.", "success")
        else:
            name = (request.form.get("name") or "").strip()
            cat = (request.form.get("category") or CAT_UNSKILLED).strip()
            if cat not in WORKER_CATEGORIES:
                cat = CAT_UNSKILLED
            if name and not Skill.query.filter_by(name=name).first():
                db.session.add(Skill(name=name, category=cat))
                db.session.commit()
                flash(f"Skill '{name}' added ({cat}).", "success")
        return redirect(url_for("admin.skills"))
    rows = Skill.query.order_by(Skill.category, Skill.name).all()
    return render_template("admin/skills.html", rows=rows, categories=WORKER_CATEGORIES)


# ---------------------------------------------------------------------------
# Rate cards
# ---------------------------------------------------------------------------
@bp.route("/rates")
@login_required
@role_required(ROLE_ADMIN)
def rates():
    rows = RateCard.query.order_by(RateCard.effective_from.desc()).all()
    return render_template("admin/rates.html", rows=rows)


@bp.route("/rates/new", methods=["GET", "POST"])
@login_required
@role_required(ROLE_ADMIN)
def rate_new():
    if request.method == "POST":
        f = request.form
        rc = RateCard(
            agency_id=int(f["agency_id"]),
            project_id=int(f["project_id"]),
            skill_id=int(f["skill_id"]),
            daily_rate=Decimal(f.get("daily_rate") or "0"),
            ot_hourly_rate=Decimal(f.get("ot_hourly_rate") or "0"),
            effective_from=parse_date(f.get("effective_from")) or date.today(),
            effective_to=parse_date(f.get("effective_to")),
        )
        db.session.add(rc)
        try:
            db.session.commit()
            flash("Rate card added.", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"Could not save: {e}", "error")
        return redirect(url_for("admin.rates"))
    return render_template(
        "admin/rate_form.html",
        agencies=Agency.query.filter_by(is_active=True).all(),
        projects=Project.query.filter_by(is_active=True).all(),
        skills=Skill.query.order_by(Skill.name).all(),
    )


# ---------------------------------------------------------------------------
# Workers (admin can also onboard)
# ---------------------------------------------------------------------------
@bp.route("/workers")
@login_required
@role_required(ROLE_ADMIN)
def workers():
    rows = Worker.query.order_by(Worker.full_name).all()
    return render_template("admin/workers.html", rows=rows)


@bp.route("/workers/new", methods=["GET", "POST"])
@login_required
@role_required(ROLE_ADMIN)
def worker_new():
    if request.method == "POST":
        f = request.form
        w = Worker(
            code=(f.get("code") or "").strip(),
            full_name=(f.get("full_name") or "").strip(),
            agency_id=int(f["agency_id"]),
            skill_id=int(f["skill_id"]) if f.get("skill_id") else None,
            gender=f.get("gender") or "",
            mobile=f.get("mobile") or "",
            aadhaar=f.get("aadhaar") or "",
            bank_name=f.get("bank_name") or "",
            account_no=f.get("account_no") or "",
            ifsc=f.get("ifsc") or "",
            onboarded_on=parse_date(f.get("onboarded_on")) or date.today(),
        )
        db.session.add(w)
        db.session.flush()

        # create a login for the worker
        if f.get("create_login"):
            uname = f.get("username") or w.code
            u = User(
                username=uname,
                display_name=w.full_name,
                role=ROLE_WORKER,
                worker_id=w.id,
                agency_id=w.agency_id,
                must_change_password=True,
            )
            u.set_password(f.get("password") or w.code)
            db.session.add(u)
        db.session.commit()
        flash(f"Worker {w.full_name} created.", "success")
        return redirect(url_for("admin.workers"))
    return render_template(
        "admin/worker_form.html",
        agencies=Agency.query.filter_by(is_active=True).all(),
        skills=Skill.query.order_by(Skill.name).all(),
    )


# ---------------------------------------------------------------------------
# NPR worker creation — auto-generates code as NPR0001YYYY, NPR0002YYYY, ...
# ---------------------------------------------------------------------------
@bp.route("/npr/new", methods=["GET", "POST"])
@login_required
@role_required(ROLE_ADMIN)
def npr_new():
    """Create a Non-Payroll worker. Auto-assigns the next NPR####YYYY code,
       puts them under the NPR vendor agency (dual-approval), and gives them
       a login with username = code, password = code, must_change_password=True."""
    agency = _ensure_npr_agency()
    year   = date.today().year
    suggested_code = _next_npr_code(year)

    if request.method == "POST":
        f = request.form
        # Honour the suggested code unless admin manually overrode (rare).
        code = (f.get("code") or suggested_code).strip()
        full_name = (f.get("full_name") or "").strip()
        if not full_name:
            flash("Worker name is required.", "error")
            return redirect(url_for("admin.npr_new"))

        # Worker
        w = Worker(
            code=code, full_name=full_name,
            agency_id=agency.id,
            skill_id=int(f["skill_id"]) if f.get("skill_id") else None,
            gender=f.get("gender") or "",
            mobile=f.get("mobile") or "",
            aadhaar=f.get("aadhaar") or "",
            bank_name=f.get("bank_name") or "",
            account_no=f.get("account_no") or "",
            ifsc=f.get("ifsc") or "",
            designation=(f.get("designation") or "").strip()[:120] or None,
            onboarded_on=parse_date(f.get("onboarded_on")) or date.today(),
        )
        db.session.add(w); db.session.flush()

        # Login — username = code, password = code (must change on first login)
        u = User(
            username=code,
            display_name=full_name,
            role=ROLE_WORKER,
            worker_id=w.id,
            agency_id=agency.id,
            must_change_password=True,
            is_active=True,
        )
        u.set_password(code)
        db.session.add(u)
        db.session.commit()

        flash(f"NPR worker '{full_name}' created with code {code}. "
              f"Login: {code} / {code} (must change on first login).", "success")
        return redirect(url_for("admin.workers"))

    return render_template(
        "admin/npr_form.html",
        suggested_code=suggested_code,
        agency=agency,
        skills=Skill.query.order_by(Skill.name).all(),
    )


@bp.route("/npr/next-code")
@login_required
@role_required(ROLE_ADMIN)
def npr_next_code():
    """JSON helper — what is the next NPR code for this year?"""
    year = int(request.args.get("year") or date.today().year)
    return jsonify(next_code=_next_npr_code(year), year=year)


# ---------------------------------------------------------------------------
# Users (light list — admin can create Procam Reps & Vendor Reps here)
# ---------------------------------------------------------------------------
@bp.route("/users")
@login_required
@role_required(ROLE_ADMIN)
def users():
    rows = User.query.order_by(User.username).all()
    return render_template("admin/users.html", rows=rows)


@bp.route("/users/new", methods=["GET", "POST"])
@login_required
@role_required(ROLE_ADMIN)
def user_new():
    if request.method == "POST":
        f = request.form
        if User.query.filter_by(username=f["username"]).first():
            flash("Username already exists.", "error")
        else:
            u = User(
                username=f["username"].strip(),
                email=f.get("email") or "",
                display_name=f.get("display_name") or f["username"],
                role=f.get("role") or ROLE_PROCAM_REP,
                agency_id=int(f["agency_id"]) if f.get("agency_id") else None,
                must_change_password=True,
            )
            u.set_password(f.get("password") or "changeme")
            db.session.add(u)
            db.session.commit()
            flash("User created.", "success")
            return redirect(url_for("admin.users"))
    return render_template(
        "admin/user_form.html",
        agencies=Agency.query.filter_by(is_active=True).all(),
    )


# ---------------------------------------------------------------------------
# Daily attendance — admin sheet
# ---------------------------------------------------------------------------
@bp.route("/attendance/daily")
@login_required
@role_required(ROLE_ADMIN)
def attendance_daily():
    """Browse + download per-day attendance."""
    from models import AttendanceApproval
    d = parse_date(request.args.get("date")) or date.today()
    rows = (Attendance.query.filter_by(work_date=d)
            .join(Worker, Worker.id == Attendance.worker_id)
            .order_by(Worker.code).all())
    return render_template("admin/attendance_daily.html", rows=rows, the_date=d)


@bp.route("/attendance/daily.xlsx")
@login_required
@role_required(ROLE_ADMIN)
def attendance_daily_xlsx():
    """One-click Excel of the day's attendance — for HR to circulate."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from flask import send_file
    import io as _io

    d = parse_date(request.args.get("date")) or date.today()
    rows = (Attendance.query.filter_by(work_date=d)
            .join(Worker, Worker.id == Attendance.worker_id)
            .order_by(Worker.code).all())

    wb = Workbook(); ws = wb.active
    ws.title = f"Attendance {d:%d-%b-%y}"[:30]
    red = PatternFill("solid", fgColor="BC1D2F")
    white_bold = Font(bold=True, color="FFFFFF")

    # Headers — Mode 2 adds Type + the self-attendance audit columns. Kiosk
    # rows simply leave the self columns blank, so the sheet works for both.
    headers = ["Code", "Worker", "Department / Vertical", "Designation", "Category",
               "Type", "IN", "OUT", "Hours",
               "Approval 1", "Approval 2", "Overall",
               "Face", "GPS", "Office", "Distance (m)", "Accuracy (m)",
               "Device", "Browser", "IP Address", "Map Link"]
    ncols = len(headers)

    # Title banner spans all columns.
    title = f"PROCAM — Daily Attendance · {d.strftime('%A, %d %B %Y')}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, color="FFFFFF", size=13)
    ws.cell(row=1, column=1).fill = red
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = white_bold; c.fill = red
        c.alignment = Alignment(horizontal="center", vertical="center")

    def _fmt_t(dt):
        # Excel cells get IST clock — matches what HR sees on the web page.
        ist = to_ist(dt)
        return ist.strftime("%H:%M:%S") if ist else ""

    def _yn(v):
        return "" if v is None else ("Yes" if v else "No")

    for i, a in enumerate(rows, start=3):
        w = a.worker
        is_self = (a.attendance_type == ATT_TYPE_SELF)
        if is_self:
            a1 = a.approval(SIDE_MANAGER); a2 = a.approval(SIDE_HR)
            vs = a1.status if a1 else "—"
            cs = a2.status if a2 else "—"
        else:
            v = a.approval("Vendor"); cap = a.approval("Client")
            vs = v.status if v else "—"
            cs = cap.status if cap else "—"
        overall = "Approved" if a.is_dual_approved else \
                  ("Declined" if "Declined" in (vs, cs) else "Pending")
        gps_txt = ("Outside" if a.outside_geofence else
                   ("" if a.gps_verified is None else ("Yes" if a.gps_verified else "No")))
        maplink = (f"https://www.google.com/maps?q={a.latitude},{a.longitude}"
                   if (a.latitude is not None and a.longitude is not None) else "")
        vals = [
            w.code if w else "",
            w.full_name if w else "",
            (w.agency.name if w and w.agency else ""),
            (w.skill.name if w and w.skill else ""),
            (w.category if w else ""),
            a.attendance_type or "Kiosk",
            _fmt_t(a.punch_in_at) or _fmt_t(a.captured_at),
            _fmt_t(a.punch_out_at),
            float(a.hours or 0),
            vs, cs, overall,
            _yn(a.face_verified) if is_self else "",
            gps_txt if is_self else "",
            a.location_name or "" if is_self else "",
            (round(a.distance_m, 1) if a.distance_m is not None else "") if is_self else "",
            (round(a.gps_accuracy, 1) if a.gps_accuracy is not None else "") if is_self else "",
            a.device or "" if is_self else "",
            a.browser or "" if is_self else "",
            a.ip_address or "" if is_self else "",
            maplink if is_self else "",
        ]
        for col, val in enumerate(vals, start=1):
            ws.cell(row=i, column=col, value=val)

    # Footer summary
    sr = len(rows) + 4
    ws.cell(row=sr, column=1, value="Summary").font = Font(bold=True)
    n_present = sum(1 for a in rows if a.status == "Present")
    n_appr = sum(1 for a in rows if a.is_dual_approved)
    n_self = sum(1 for a in rows if a.attendance_type == ATT_TYPE_SELF)
    ws.cell(row=sr + 1, column=1, value="Rows total").font = Font(bold=True)
    ws.cell(row=sr + 1, column=2, value=len(rows))
    ws.cell(row=sr + 2, column=1, value="Present").font = Font(bold=True)
    ws.cell(row=sr + 2, column=2, value=n_present)
    ws.cell(row=sr + 3, column=1, value="Fully approved (billable)").font = Font(bold=True)
    ws.cell(row=sr + 3, column=2, value=n_appr)
    ws.cell(row=sr + 4, column=1, value="Self-attendance rows").font = Font(bold=True)
    ws.cell(row=sr + 4, column=2, value=n_self)

    widths = [12, 26, 20, 18, 12, 9, 10, 10, 7, 12, 12, 11,
              7, 9, 16, 12, 12, 14, 16, 15, 30]
    for i, wd in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = wd

    buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
    fn = f"Procam_Attendance_{d.strftime('%Y-%m-%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fn,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ---------------------------------------------------------------------------
# HR Approval queue (for hr_only agencies — i.e. Procam in-house)
# ---------------------------------------------------------------------------
@bp.route("/attendance/approvals", methods=["GET", "POST"])
@login_required
@role_required(ROLE_ADMIN)
def attendance_approvals():
    """HR approves office-employee punches (Client side). Pending only."""
    from datetime import datetime
    from models import AttendanceApproval, Agency, APPROVER_HR_ONLY

    if request.method == "POST":
        action = request.form.get("action")
        att_ids = request.form.getlist("att_id", type=int)
        if not att_ids:
            flash("Pick at least one row.", "warning")
            return redirect(url_for("admin.attendance_approvals"))

        new_status = "Approved" if action == "approve" else \
                     "Declined" if action == "decline" else None
        if new_status:
            for aid in att_ids:
                att = Attendance.query.get(aid)
                if not att:
                    continue
                ap = att.approval("Client")
                if not ap:
                    ap = AttendanceApproval(attendance_id=att.id, side="Client")
                    db.session.add(ap)
                ap.status = new_status
                ap.decided_by_id = current_user.id
                ap.decided_at = datetime.utcnow()
            db.session.commit()
            flash(f"{len(att_ids)} punch(es) {new_status.lower()}.", "success")
        return redirect(url_for("admin.attendance_approvals"))

    d = parse_date(request.args.get("date")) or date.today()
    # Pull rows where the worker's agency is HR-only AND the Client approval is Pending/missing
    hr_agencies = [a.id for a in Agency.query.filter_by(approver_mode=APPROVER_HR_ONLY).all()]
    q = (Attendance.query
         .join(Worker, Worker.id == Attendance.worker_id)
         .filter(Worker.agency_id.in_(hr_agencies),
                 Attendance.work_date == d))
    pending = [a for a in q.order_by(Worker.code).all()
               if not (a.approval("Client") and a.approval("Client").status == "Approved")]
    decided_today = [a for a in q.all() if a.is_dual_approved]
    return render_template("admin/attendance_approvals.html",
                           pending=pending, decided=decided_today, the_date=d)


# ---------------------------------------------------------------------------
# Browser-upload PRERNA importer — for Render where Terminal access is painful.
# Logs in as HR Admin → /admin/import-employees → drag & drop PRERNA xlsx →
# all 124 employees created/updated with correct PRERNA Role mapping,
# Tanima Mukherjee (or whoever PRERNA flags as HR_ADMIN) auto-promoted to Admin.
# ---------------------------------------------------------------------------
_PRERNA_ROLE_MAP = {
    "SUPER_ADMIN": ROLE_ADMIN,
    "HR_ADMIN":    ROLE_ADMIN,
    "MANAGER":     ROLE_PROCAM_REP,
    "EMPLOYEE":    ROLE_WORKER,
}


def _ensure_procam_agency() -> Agency:
    """In-house PROCAM agency, mode = none (no approval)."""
    a = Agency.query.filter_by(code="PROCAM").first()
    if a:
        if a.approver_mode != APPROVER_NONE:
            a.approver_mode = APPROVER_NONE
        return a
    a = Agency(
        code="PROCAM", name="Procam Logistics (In-house)",
        contact_person="HR Admin", email="hr@procamlogistics.com",
        address="731, Westend Mall, District Centre, Janakpuri, New Delhi-110058.",
        gstin=None, pan=None,
        default_gst_rate=Decimal("0.00"), default_tds_rate=Decimal("0.00"),
        onboarded_on=date.today(), is_active=True,
        approver_mode=APPROVER_NONE,
    )
    db.session.add(a); db.session.flush()
    return a


def _ensure_skill_inline(name: str) -> Skill:
    name = (name or "Employee").strip()[:80] or "Employee"
    sk = Skill.query.filter_by(name=name).first()
    if sk:
        return sk
    sk = Skill(name=name, category=CAT_SKILLED)
    db.session.add(sk); db.session.flush()
    return sk


# ---------------------------------------------------------------------------
# Face enrolment audit — who has enrolled, when, and what the photo looks like
# ---------------------------------------------------------------------------
@bp.route("/face-enrolments")
@login_required
@role_required(ROLE_ADMIN)
def face_enrolments():
    """Grid + stats. Shows every worker, whether they've enrolled, when, how
       many poses, and the snapshot photo captured during their Centre pose."""
    from models import FaceTemplate
    workers = Worker.query.filter_by(is_active=True).order_by(Worker.full_name).all()
    enrolled_ids = {t.worker_id: t for t in FaceTemplate.query.all()}
    total_workers   = len(workers)
    enrolled_count  = sum(1 for w in workers if w.id in enrolled_ids)
    pending_count   = total_workers - enrolled_count
    has_snapshot    = sum(1 for t in enrolled_ids.values() if t.snapshot_b64)
    return render_template("admin/face_enrolments.html",
                           workers=workers, templates=enrolled_ids,
                           total=total_workers, enrolled=enrolled_count,
                           pending=pending_count, has_snapshot=has_snapshot)


@bp.route("/face-enrolments/reset/<int:worker_id>", methods=["POST"])
@login_required
@role_required(ROLE_ADMIN)
def face_reset_one(worker_id: int):
    """Wipe one worker's face template — they'll be forced to re-enrol on next login."""
    from models import FaceTemplate
    tpl = FaceTemplate.query.filter_by(worker_id=worker_id).first()
    w   = Worker.query.get(worker_id)
    if tpl:
        db.session.delete(tpl)
        db.session.commit()
        flash(f"Face wiped for {w.full_name if w else worker_id}. "
              f"They'll be forced to re-enrol on next login.", "success")
    else:
        flash("No face template to wipe for that worker.", "warning")
    return redirect(url_for("admin.face_enrolments"))


@bp.route("/import-employees", methods=["GET", "POST"])
@login_required
@role_required(ROLE_ADMIN)
def import_employees():
    """Browser-based PRERNA xlsx importer. POST an xlsx file → bootstrap
       everyone in one click. Idempotent — safe to re-run on updates."""
    if request.method == "GET":
        # Show a small upload form with the current employee count
        return render_template("admin/import_employees.html",
                               current_count=User.query.count(),
                               workers=Worker.query.count())

    # POST — handle the upload
    f = request.files.get("file")
    if not f or not f.filename.lower().endswith((".xlsx", ".xlsm")):
        flash("Please choose a PRERNA xlsx file.", "error")
        return redirect(url_for("admin.import_employees"))

    # Save to a tempfile so openpyxl can read it
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tf:
        f.save(tf.name)
        tmp_path = tf.name

    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename=tmp_path, read_only=True, data_only=True)

        ws = None
        for name in wb.sheetnames:
            if "employee" in name.lower() or "master" in name.lower():
                ws = wb[name]; break
        if ws is None:
            ws = wb.active

        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h else "" for h in next(rows_iter)]
        idx = {h.lower(): i for i, h in enumerate(headers)}

        def col(row, key):
            i = idx.get(key.lower())
            return row[i] if i is not None and i < len(row) else None

        agency = _ensure_procam_agency()
        created_w = updated_w = created_u = updated_u = 0

        for row in rows_iter:
            if not row or all(v in (None, "") for v in row):
                continue
            emp_code = col(row, "Emp Code")
            full_name = col(row, "Full Name")
            if not emp_code or not full_name:
                continue
            emp_code  = str(emp_code).strip()
            full_name = str(full_name).strip()

            vertical    = (str(col(row, "Vertical")    or "").strip())[:80] or None
            designation = (str(col(row, "Designation") or "").strip())[:120] or "Employee"
            grade       = (str(col(row, "Grade")       or "").strip())[:20] or None
            mgr_code    = (str(col(row, "Mgr Code")    or "").strip())[:32] or None
            prerna_role = (str(col(row, "Role")        or "").strip()) or "EMPLOYEE"
            app_role    = _PRERNA_ROLE_MAP.get(
                prerna_role.strip().upper().replace(" ", "_"), ROLE_WORKER)

            sk = _ensure_skill_inline(designation)

            # Worker
            w = Worker.query.filter_by(code=emp_code).first()
            if not w:
                w = Worker(
                    code=emp_code, full_name=full_name,
                    agency_id=agency.id, skill_id=sk.id,
                    is_active=True, onboarded_on=date.today(),
                    manager_code=mgr_code, designation=designation,
                    vertical=vertical, grade=grade,
                )
                db.session.add(w); db.session.flush()
                created_w += 1
            else:
                w.full_name = full_name
                w.agency_id = agency.id
                w.skill_id  = sk.id
                w.is_active = True
                w.manager_code = mgr_code
                w.designation  = designation
                w.vertical     = vertical
                w.grade        = grade
                updated_w += 1

            # User
            u = User.query.filter_by(username=emp_code).first()
            if not u:
                u = User(
                    username=emp_code, display_name=full_name,
                    role=app_role, is_active=True,
                    must_change_password=True,
                    worker_id=w.id,
                )
                u.set_password(emp_code)
                db.session.add(u)
                created_u += 1
            else:
                u.display_name = full_name
                u.worker_id    = w.id
                u.role         = app_role          # re-classify on every upload
                u.is_active    = True
                updated_u += 1

            if (created_w + updated_w) % 50 == 0:
                db.session.commit()

        db.session.commit()
        wb.close()

        # Tally final role breakdown for the success page
        role_counts = Counter(u.role for u in User.query.all())

        flash(f"Imported successfully. Workers: {created_w} new + {updated_w} updated. "
              f"Users: {created_u} new + {updated_u} updated. "
              f"Roles now: {dict(role_counts)}", "success")
        return redirect(url_for("admin.import_employees"))

    except Exception as e:
        db.session.rollback()
        flash(f"Import failed: {type(e).__name__}: {e}", "error")
        return redirect(url_for("admin.import_employees"))
    finally:
        try: os.unlink(tmp_path)
        except OSError: pass


# ===========================================================================
# Employee Self Attendance (Mode 2) — admin configuration
# ===========================================================================
def _chk(f, name: str) -> bool:
    """Checkbox helper — HTML checkboxes only POST when ticked."""
    return f.get(name) in ("on", "1", "true", "yes")


def _int_or(f, name: str, default):
    try:
        v = (f.get(name) or "").strip()
        return int(v) if v != "" else default
    except (TypeError, ValueError):
        return default


def _float_or_none(f, name: str):
    v = (f.get(name) or "").strip()
    if v == "":
        return None
    try:
        return float(v)
    except ValueError:
        return None


@bp.route("/self-attendance/settings", methods=["GET", "POST"])
@login_required
@role_required(ROLE_ADMIN)
def self_attendance_settings():
    """The 'Self Attendance Settings' admin panel."""
    s = SelfAttendanceSettings.get()
    if request.method == "POST":
        f = request.form
        s.enable_self_attendance   = _chk(f, "enable_self_attendance")
        s.enable_face_verification = _chk(f, "enable_face_verification")
        s.enable_gps               = _chk(f, "enable_gps")
        s.enable_geofence          = _chk(f, "enable_geofence")
        s.allow_outside_radius     = _chk(f, "allow_outside_radius")
        s.require_live_camera      = _chk(f, "require_live_camera")
        s.capture_photo            = _chk(f, "capture_photo")
        s.max_gps_accuracy_m       = _int_or(f, "max_gps_accuracy_m", 50)
        s.cooldown_seconds         = _int_or(f, "cooldown_seconds", 60)
        s.window_start             = (f.get("window_start") or "").strip() or None
        s.window_end               = (f.get("window_end") or "").strip() or None
        s.late_mark_after          = (f.get("late_mark_after") or "").strip() or None
        s.default_office_latitude  = _float_or_none(f, "default_office_latitude")
        s.default_office_longitude = _float_or_none(f, "default_office_longitude")
        s.default_office_radius_m  = _int_or(f, "default_office_radius_m", 100)
        db.session.commit()
        flash("Self attendance settings saved.", "success")
        return redirect(url_for("admin.self_attendance_settings"))
    offices = Office.query.order_by(Office.name).all()
    return render_template("admin/self_attendance_settings.html",
                           s=s, offices=offices)


# --------------------------- Offices CRUD ----------------------------------
@bp.route("/offices")
@login_required
@role_required(ROLE_ADMIN)
def offices():
    rows = Office.query.order_by(Office.name).all()
    return render_template("admin/offices.html", rows=rows)


@bp.route("/offices/new", methods=["GET", "POST"])
@login_required
@role_required(ROLE_ADMIN)
def office_new():
    s = SelfAttendanceSettings.get()
    if request.method == "POST":
        f = request.form
        o = Office(
            code=(f.get("code") or "").strip().upper(),
            name=(f.get("name") or "").strip(),
            city=(f.get("city") or "").strip() or None,
            address=(f.get("address") or "").strip() or None,
            latitude=_float_or_none(f, "latitude"),
            longitude=_float_or_none(f, "longitude"),
            radius_m=_int_or(f, "radius_m", s.default_office_radius_m or 100),
            is_active=_chk(f, "is_active"),
        )
        db.session.add(o)
        db.session.commit()
        flash(f"Office '{o.name}' added.", "success")
        return redirect(url_for("admin.offices"))
    return render_template("admin/office_form.html", o=None, s=s,
                           eligible_workers=[])


@bp.route("/offices/<int:oid>/edit", methods=["GET", "POST"])
@login_required
@role_required(ROLE_ADMIN)
def office_edit(oid):
    o = Office.query.get_or_404(oid)
    s = SelfAttendanceSettings.get()
    if request.method == "POST":
        f = request.form
        o.code = (f.get("code") or "").strip().upper()
        o.name = (f.get("name") or "").strip()
        o.city = (f.get("city") or "").strip() or None
        o.address = (f.get("address") or "").strip() or None
        o.latitude = _float_or_none(f, "latitude")
        o.longitude = _float_or_none(f, "longitude")
        o.radius_m = _int_or(f, "radius_m", 100)
        o.is_active = _chk(f, "is_active")
        # Update the employee <-> office assignment from the multi-select.
        chosen_ids = set(request.form.getlist("worker_ids", type=int))
        o.workers = Worker.query.filter(Worker.id.in_(chosen_ids)).all() \
            if chosen_ids else []
        db.session.commit()
        flash(f"Office '{o.name}' updated.", "success")
        return redirect(url_for("admin.offices"))
    # Only office staff (in-house PROCAM agency) can be geofence-assigned.
    in_house = [a.id for a in Agency.query.filter_by(approver_mode=APPROVER_NONE).all()]
    eligible_workers = (Worker.query
                        .filter(Worker.is_active.is_(True),
                                Worker.agency_id.in_(in_house))
                        .order_by(Worker.full_name).all()) if in_house else []
    assigned_ids = {w.id for w in o.workers}
    return render_template("admin/office_form.html", o=o, s=s,
                           eligible_workers=eligible_workers,
                           assigned_ids=assigned_ids)


# ------------------- HR final approval of self attendance ------------------
@bp.route("/self-attendance/approvals", methods=["GET", "POST"])
@login_required
@role_required(ROLE_ADMIN)
def self_attendance_approvals():
    """HR gives the FINAL sign-off on employee self attendance. A punch appears
       here once its Manager step is Approved (or when the employee has no
       manager on file, so there is no Manager step)."""
    d = parse_date(request.args.get("date")) or date.today()

    if request.method == "POST":
        from datetime import datetime as _dt
        action = request.form.get("action")
        att_ids = request.form.getlist("att_id", type=int)
        new_status = "Approved" if action == "approve" else \
                     "Declined" if action == "decline" else None
        if new_status and att_ids:
            for aid in att_ids:
                att = Attendance.query.get(aid)
                if not att or att.attendance_type != ATT_TYPE_SELF:
                    continue
                ap = att.approval(SIDE_HR)
                if not ap:
                    ap = AttendanceApproval(attendance_id=att.id, side=SIDE_HR)
                    db.session.add(ap)
                ap.status = new_status
                ap.decided_by_id = current_user.id
                ap.decided_at = _dt.utcnow()
            db.session.commit()
            flash(f"{len(att_ids)} self punch(es) {new_status.lower()} by HR.", "success")
        return redirect(url_for("admin.self_attendance_approvals", date=d.isoformat()))

    rows = (Attendance.query
            .join(Worker, Worker.id == Attendance.worker_id)
            .filter(Attendance.attendance_type == ATT_TYPE_SELF,
                    Attendance.work_date == d)
            .order_by(Worker.code).all())
    pending, decided = [], []
    for a in rows:
        mgr = a.approval(SIDE_MANAGER)
        hr = a.approval(SIDE_HR)
        manager_ok = (mgr is None) or (mgr.status == "Approved")
        hr_done = hr and hr.status in ("Approved", "Declined")
        # Ready for HR only once the manager step is cleared and HR hasn't decided.
        if manager_ok and not hr_done:
            pending.append(a)
        elif hr_done:
            decided.append(a)
    return render_template("admin/self_attendance_approvals.html",
                           pending=pending, decided=decided, the_date=d)
