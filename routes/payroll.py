"""NPR Worker Payroll — per-worker payslips from approved attendance.

Mirrors the invoice generator but pivots PER WORKER instead of per agency.
The same source-of-truth rows (dual-approved Attendance) drive both
documents, so the invoice Procam pays the agency always reconciles to the
sum of the payslips the agency disburses to each worker.

Workflow:
  1. Admin picks (agency, project, period_start, period_end).
  2. For every worker under that agency: sum dual-approved billable days
     in the window, resolve the rate from the active RateCard, compute
     gross. (Statutory deductions left at zero for now — agencies handle
     PF/ESI internally; the field is on the model for future use.)
  3. Procam HR generates + locks the payslip; the agency disburses.

NOTE on terminology: "NPR" = non-payroll resource = contract worker under
the manpower vendor. Procam computes the bill so the worker has an
auditable record from the principal employer.
"""
from datetime import date, datetime
from decimal import Decimal

from flask import (Blueprint, render_template, request, redirect, url_for,
                   flash, send_file, abort)
from flask_login import login_required, current_user

from models import (db, Agency, Project, Worker, Attendance,
                   AttendanceApproval, ATT_PRESENT, ATT_HALF, ATT_REGULARIZED,
                   ROLE_ADMIN)
from utils import role_required, parse_date, indian_commas, amount_in_words

bp = Blueprint("payroll", __name__)


# ---------------------------------------------------------------------------
def _worker_billable_days(worker_id: int, project_id: int,
                          period_start: date, period_end: date):
    """Return (present_full, half_days, list_of_billable_dates) for the worker
       in the given window. Only dual-approved attendance counts."""
    rows = (Attendance.query
            .filter(Attendance.worker_id == worker_id,
                    Attendance.project_id == project_id,
                    Attendance.work_date >= period_start,
                    Attendance.work_date <= period_end,
                    Attendance.status.in_([ATT_PRESENT, ATT_HALF, ATT_REGULARIZED]))
            .all())
    present = 0; half = 0; dates = []
    for a in rows:
        if not a.is_dual_approved:
            continue
        if a.status == ATT_HALF:
            half += 1
        else:
            present += 1
        dates.append((a.work_date, a.status))
    return present, half, dates


def _resolve_rate(agency_id, project_id, skill_id, on_day):
    from models import RateCard
    q = (RateCard.query
         .filter(RateCard.agency_id == agency_id,
                 RateCard.project_id == project_id,
                 RateCard.skill_id == skill_id,
                 RateCard.effective_from <= on_day)
         .order_by(RateCard.effective_from.desc()))
    for rc in q.all():
        if rc.effective_to is None or rc.effective_to >= on_day:
            return rc
    return None


def compute_worker_pay(worker: Worker, project_id: int,
                       period_start: date, period_end: date) -> dict:
    """Returns a dict with every line item for one worker × project × period."""
    present, half, dates = _worker_billable_days(
        worker.id, project_id, period_start, period_end)
    if not dates:
        return {
            "worker": worker, "present_days": 0, "half_days": 0,
            "billable_days": Decimal("0"), "daily_rate": Decimal("0"),
            "gross": Decimal("0"), "dates": [],
        }
    # Resolve rate per day (catches mid-period rate changes correctly)
    total = Decimal("0")
    last_rate = Decimal("0")
    for d, status in dates:
        rc = _resolve_rate(worker.agency_id, project_id, worker.skill_id, d)
        rate = (rc.daily_rate if rc else Decimal("0"))
        last_rate = rate
        if status == ATT_HALF:
            total += rate * Decimal("0.5")
        else:
            total += rate
    billable_days = Decimal(present) + (Decimal(half) * Decimal("0.5"))
    return {
        "worker": worker, "present_days": present, "half_days": half,
        "billable_days": billable_days, "daily_rate": last_rate,
        "gross": total, "dates": dates,
    }


# ---------------------------------------------------------------------------
# Admin pages
# ---------------------------------------------------------------------------
@bp.route("/payroll", methods=["GET", "POST"])
@login_required
@role_required(ROLE_ADMIN)
def payroll_run():
    """Generate the per-worker payslip preview for a period."""
    agencies = Agency.query.order_by(Agency.name).all()
    projects = Project.query.order_by(Project.name).all()
    result = None

    if request.method == "POST":
        agency_id = int(request.form.get("agency_id") or 0)
        project_id = int(request.form.get("project_id") or 0)
        ps = parse_date(request.form.get("period_start"))
        pe = parse_date(request.form.get("period_end"))
        if not (agency_id and project_id and ps and pe):
            flash("Pick agency, project, and a valid period.", "warning")
            return redirect(url_for("payroll.payroll_run"))

        agency = Agency.query.get_or_404(agency_id)
        project = Project.query.get_or_404(project_id)
        workers = Worker.query.filter_by(agency_id=agency_id, is_active=True).all()
        lines = []
        grand_total = Decimal("0")
        for w in workers:
            line = compute_worker_pay(w, project_id, ps, pe)
            if line["gross"] > 0:
                lines.append(line)
                grand_total += line["gross"]

        result = {
            "agency": agency, "project": project,
            "period_start": ps, "period_end": pe,
            "lines": lines, "grand_total": grand_total,
        }

    return render_template("payroll/run.html",
                           agencies=agencies, projects=projects, result=result)


@bp.route("/payroll/worker/<int:worker_id>")
@login_required
@role_required(ROLE_ADMIN)
def worker_payslip(worker_id):
    """Printable per-worker payslip (Procam branded). URL params: project_id, ps, pe."""
    worker = Worker.query.get_or_404(worker_id)
    project_id = int(request.args.get("project_id") or 0)
    ps = parse_date(request.args.get("ps"))
    pe = parse_date(request.args.get("pe"))
    if not (project_id and ps and pe):
        abort(400)
    project = Project.query.get_or_404(project_id)
    line = compute_worker_pay(worker, project_id, ps, pe)
    return render_template("payroll/payslip_print.html",
                           worker=worker, project=project, line=line,
                           period_start=ps, period_end=pe,
                           agency=worker.agency,
                           generated_at=datetime.utcnow())


@bp.route("/payroll/export.xlsx")
@login_required
@role_required(ROLE_ADMIN)
def export_xlsx():
    """One Excel for the agency: every worker's payslip summary in one sheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    import io as _io

    agency_id = int(request.args.get("agency_id") or 0)
    project_id = int(request.args.get("project_id") or 0)
    ps = parse_date(request.args.get("ps"))
    pe = parse_date(request.args.get("pe"))
    if not (agency_id and project_id and ps and pe):
        abort(400)
    agency = Agency.query.get_or_404(agency_id)
    project = Project.query.get_or_404(project_id)

    wb = Workbook(); ws = wb.active
    ws.title = f"Payslips {ps:%b}-{pe:%y}"[:30]
    hdr_fill = PatternFill("solid", fgColor="BC1D2F")
    hdr_font = Font(bold=True, color="FFFFFF")

    title = f"Procam — Worker Payslip Summary · {agency.name} · {project.name} · {ps:%d-%b-%Y} → {pe:%d-%b-%Y}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, color="FFFFFF", size=12)
    ws.cell(row=1, column=1).fill = hdr_fill
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    headers = ["Code", "Worker", "Skill", "Category", "Present", "Half", "Billable Days", "Daily Rate ₹", "Gross ₹"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")

    workers = Worker.query.filter_by(agency_id=agency_id, is_active=True).all()
    r = 3
    grand = Decimal("0")
    for w in workers:
        line = compute_worker_pay(w, project_id, ps, pe)
        if line["gross"] <= 0:
            continue
        ws.cell(row=r, column=1, value=w.code)
        ws.cell(row=r, column=2, value=w.full_name)
        ws.cell(row=r, column=3, value=w.skill.name if w.skill else "")
        ws.cell(row=r, column=4, value=w.category)
        ws.cell(row=r, column=5, value=line["present_days"])
        ws.cell(row=r, column=6, value=line["half_days"])
        ws.cell(row=r, column=7, value=float(line["billable_days"]))
        ws.cell(row=r, column=8, value=float(line["daily_rate"]))
        ws.cell(row=r, column=9, value=float(line["gross"]))
        grand += line["gross"]
        r += 1
    # Total row
    ws.cell(row=r, column=8, value="GRAND TOTAL").font = Font(bold=True)
    ws.cell(row=r, column=9, value=float(grand)).font = Font(bold=True)

    for i in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + i)].width = 14
    ws.column_dimensions["B"].width = 28

    buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
    fn = f"Worker_Payslips_{agency.code}_{ps:%Y%m%d}_to_{pe:%Y%m%d}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fn,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
