"""Leave module — any logged-in employee can apply for leave.

Two-step approval (mirrors self-attendance): the employee's line **Manager**
approves first, then **HR** gives the final sign-off. When an employee has no
manager on file, the manager step is marked 'NA' and only HR approval is needed.

Leave types only (Casual / Sick / Earned / Unpaid) — no balance tracking in
this phase. Nothing here touches attendance capture (kiosk or self).
"""
from datetime import date, datetime
from decimal import Decimal

from flask import (Blueprint, render_template, request, redirect, url_for,
                   flash, abort)
from flask_login import login_required, current_user

from models import (db, Worker, User, LeaveRequest, LEAVE_TYPES,
                    LV_PENDING, LV_APPROVED, LV_DECLINED, LV_NA,
                    ROLE_ADMIN, ROLE_PROCAM_REP)
from utils import role_required, parse_date

bp = Blueprint("leave", __name__)


def _leave_days(from_d: date, to_d: date, half_day: bool) -> Decimal:
    """Inclusive day count; a half-day is always 0.5 (single date)."""
    if half_day:
        return Decimal("0.5")
    delta = (to_d - from_d).days + 1
    return Decimal(str(max(delta, 1)))


# ---------------------------------------------------------------------------
# Apply for leave — available to EVERY logged-in employee
# ---------------------------------------------------------------------------
@bp.route("/apply", methods=["GET", "POST"])
@login_required
def apply():
    # Admin / managers may file on behalf of any employee; a normal employee
    # files only for themselves.
    can_pick = current_user.role in (ROLE_ADMIN, ROLE_PROCAM_REP)

    if request.method == "POST":
        f = request.form
        # Resolve the target employee.
        if can_pick and f.get("worker_id"):
            worker = Worker.query.get(int(f["worker_id"]))
        else:
            worker = current_user.worker
        if not worker:
            flash("No employee record linked to this request.", "error")
            return redirect(url_for("leave.apply"))

        leave_type = f.get("leave_type") or "Casual"
        if leave_type not in LEAVE_TYPES:
            leave_type = "Casual"
        half_day = f.get("half_day") in ("on", "1", "true", "yes")
        from_d = parse_date(f.get("from_date"))
        to_d = parse_date(f.get("to_date")) or from_d
        if not from_d:
            flash("Please pick a start date.", "error")
            return redirect(url_for("leave.apply"))
        if half_day:
            to_d = from_d          # a half-day is a single date
        if to_d < from_d:
            flash("End date can't be before the start date.", "error")
            return redirect(url_for("leave.apply"))

        lr = LeaveRequest(
            worker_id=worker.id,
            leave_type=leave_type,
            from_date=from_d, to_date=to_d,
            half_day=half_day,
            days=_leave_days(from_d, to_d, half_day),
            reason=(f.get("reason") or "").strip(),
            requested_by_id=current_user.id,
            # No line manager -> skip the manager step entirely.
            manager_status=(LV_PENDING if (worker.manager_code or "").strip() else LV_NA),
        )
        lr.recompute_status()
        db.session.add(lr)
        db.session.commit()
        flash("Leave request submitted.", "success")
        return redirect(url_for("leave.apply"))

    workers = None
    if can_pick:
        workers = Worker.query.filter_by(is_active=True).order_by(Worker.full_name).all()
    mine = []
    if current_user.worker_id:
        mine = (LeaveRequest.query
                .filter_by(worker_id=current_user.worker_id)
                .order_by(LeaveRequest.created_at.desc()).limit(20).all())
    return render_template("leave/apply.html", leave_types=LEAVE_TYPES,
                           workers=workers, mine=mine, can_pick=can_pick)


# ---------------------------------------------------------------------------
# Shared filtered query for the register + export
# ---------------------------------------------------------------------------
def _leave_query():
    status = (request.args.get("status") or "All").strip()
    ltype = (request.args.get("type") or "All").strip()
    q_text = (request.args.get("q") or "").strip()
    d_from = parse_date(request.args.get("from"))
    d_to = parse_date(request.args.get("to"))

    q = LeaveRequest.query.join(Worker, Worker.id == LeaveRequest.worker_id)
    if status and status != "All":
        q = q.filter(LeaveRequest.status == status)
    if ltype and ltype != "All":
        q = q.filter(LeaveRequest.leave_type == ltype)
    if q_text:
        like = f"%{q_text}%"
        q = q.filter(db.or_(Worker.full_name.ilike(like), Worker.code.ilike(like)))
    if d_from:
        q = q.filter(LeaveRequest.to_date >= d_from)
    if d_to:
        q = q.filter(LeaveRequest.from_date <= d_to)
    q = q.order_by(LeaveRequest.created_at.desc())
    filters = {"status": status, "type": ltype, "q": q_text,
               "from": request.args.get("from") or "", "to": request.args.get("to") or ""}
    return q, filters


@bp.route("/register")
@login_required
@role_required(ROLE_PROCAM_REP, ROLE_ADMIN)
def register():
    """Full leave register — all employees, all statuses, filterable."""
    q, filters = _leave_query()
    rows = q.all()
    counts = {
        "Pending": LeaveRequest.query.filter_by(status=LV_PENDING).count(),
        "Approved": LeaveRequest.query.filter_by(status=LV_APPROVED).count(),
        "Declined": LeaveRequest.query.filter_by(status=LV_DECLINED).count(),
    }
    counts["Total"] = counts["Pending"] + counts["Approved"] + counts["Declined"]
    return render_template("leave/register.html", rows=rows, filters=filters,
                           counts=counts, leave_types=LEAVE_TYPES, shown=len(rows))


@bp.route("/register.xlsx")
@login_required
@role_required(ROLE_PROCAM_REP, ROLE_ADMIN)
def register_export():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from flask import send_file
    import io as _io
    from utils import to_ist

    q, _ = _leave_query()
    rows = q.all()

    wb = Workbook(); ws = wb.active
    ws.title = "Leave"
    red = PatternFill("solid", fgColor="BC1D2F")
    white_bold = Font(bold=True, color="FFFFFF")
    headers = ["Code", "Employee", "Type", "From", "To", "Days", "Half-day",
               "Status", "Manager", "HR", "Reason", "Applied On"]
    ncols = len(headers)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=1, column=1, value=f"PROCAM — Leave Register ({len(rows)})").font = \
        Font(bold=True, color="FFFFFF", size=13)
    ws.cell(row=1, column=1).fill = red
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = white_bold; c.fill = red
        c.alignment = Alignment(horizontal="center", vertical="center")

    for i, r in enumerate(rows, start=3):
        w = r.worker
        applied = to_ist(r.created_at)
        vals = [
            w.code if w else "", w.full_name if w else "",
            r.leave_type,
            r.from_date.strftime("%d-%b-%Y") if r.from_date else "",
            r.to_date.strftime("%d-%b-%Y") if r.to_date else "",
            float(r.days or 0), "Yes" if r.half_day else "",
            r.status,
            r.manager_status, r.hr_status,
            r.reason or "",
            applied.strftime("%d-%b-%Y %H:%M") if applied else "",
        ]
        for col, val in enumerate(vals, start=1):
            ws.cell(row=i, column=col, value=val)

    widths = [12, 26, 10, 12, 12, 7, 9, 11, 11, 11, 34, 18]
    for i, wd in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = wd
    buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
    fn = f"Procam_Leave_{date.today():%Y-%m-%d}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fn,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ---------------------------------------------------------------------------
# Approvals — Manager step, then HR step
# ---------------------------------------------------------------------------
def _apply_decision(rows_ids, stage, decision):
    """Apply an approve/decline to a batch of leave rows for one stage."""
    my_code = current_user.worker.code if current_user.worker else None
    n = 0
    for lid in rows_ids:
        lr = LeaveRequest.query.get(lid)
        if not lr:
            continue
        if stage == "manager":
            # Managers may only decide their own reports; Admin may decide any.
            if current_user.role != ROLE_ADMIN:
                w = lr.worker
                if not (w and my_code and w.manager_code == my_code):
                    continue
            lr.manager_status = decision
            lr.manager_by_id = current_user.id
            lr.manager_at = datetime.utcnow()
        else:  # hr
            lr.hr_status = decision
            lr.hr_by_id = current_user.id
            lr.hr_at = datetime.utcnow()
        lr.recompute_status()
        n += 1
    db.session.commit()
    return n


@bp.route("/approvals", methods=["GET", "POST"])
@login_required
def manager_approvals():
    """Manager step — approve/decline leave for direct reports."""
    if current_user.role not in (ROLE_PROCAM_REP, ROLE_ADMIN):
        abort(403)
    my_code = current_user.worker.code if current_user.worker else None

    if request.method == "POST":
        action = request.form.get("action")
        ids = request.form.getlist("leave_id", type=int)
        decision = LV_APPROVED if action == "approve" else LV_DECLINED if action == "decline" else None
        if decision and ids:
            n = _apply_decision(ids, "manager", decision)
            flash(f"{n} leave request(s) {decision.lower()} (manager step).", "success")
        return redirect(url_for("leave.manager_approvals"))

    q = (LeaveRequest.query
         .join(Worker, Worker.id == LeaveRequest.worker_id)
         .filter(LeaveRequest.manager_status == LV_PENDING))
    if current_user.role != ROLE_ADMIN:
        q = q.filter(Worker.manager_code == my_code)
    pending = q.order_by(LeaveRequest.created_at.desc()).all()
    return render_template("leave/approvals.html", stage="manager",
                           pending=pending)


@bp.route("/hr-approvals", methods=["GET", "POST"])
@login_required
@role_required(ROLE_ADMIN)
def hr_approvals():
    """HR step — final sign-off once the manager step is cleared (or NA)."""
    if request.method == "POST":
        action = request.form.get("action")
        ids = request.form.getlist("leave_id", type=int)
        decision = LV_APPROVED if action == "approve" else LV_DECLINED if action == "decline" else None
        if decision and ids:
            n = _apply_decision(ids, "hr", decision)
            flash(f"{n} leave request(s) {decision.lower()} by HR.", "success")
        return redirect(url_for("leave.hr_approvals"))

    # Ready for HR: manager step cleared (Approved or NA) and HR still Pending.
    rows = (LeaveRequest.query
            .filter(LeaveRequest.manager_status.in_([LV_APPROVED, LV_NA]),
                    LeaveRequest.hr_status == LV_PENDING)
            .order_by(LeaveRequest.created_at.desc()).all())
    return render_template("leave/approvals.html", stage="hr", pending=rows)
