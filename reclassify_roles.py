"""Re-classify the 124 PROCAM employees you already imported, based on the
   PRERNA Role column. Use this when import_prerna.py was run BEFORE the
   role-mapping fix and everyone is currently stuck as ROLE_WORKER.

   Reads the same xlsx, finds each User by emp_code, and updates their role.

   Usage (from project root, with .venv active):
       python reclassify_roles.py "/Users/nileshsinha/Desktop/PRERNA_Full_Backup_2026-04-09 1.48.14 PM.xlsx"

   Idempotent — safe to run repeatedly.
"""
import sys, os, glob
from collections import Counter
from openpyxl import load_workbook

from app import create_app
from models import db, User, ROLE_ADMIN, ROLE_PROCAM_REP, ROLE_WORKER

PRERNA_ROLE_MAP = {
    "SUPER_ADMIN": ROLE_ADMIN,
    "HR_ADMIN":    ROLE_ADMIN,
    "MANAGER":     ROLE_PROCAM_REP,
    "EMPLOYEE":    ROLE_WORKER,
}


def _role_for(prerna_role: str) -> str:
    key = (prerna_role or "").strip().upper().replace(" ", "_")
    return PRERNA_ROLE_MAP.get(key, ROLE_WORKER)


def _resolve_xlsx(path: str) -> str:
    """If the literal path exists, use it. Otherwise glob the same directory
       for a PRERNA*.xlsx — handles the U+202F narrow no-break space in the
       filename that prevents typed paths from matching."""
    if os.path.isfile(path):
        return path
    dir_ = os.path.dirname(path) or "."
    base = os.path.basename(path)
    # Try a relaxed match: strip everything after 'PRERNA_Full_Backup_'
    candidates = glob.glob(os.path.join(dir_, "PRERNA_Full_Backup_*.xlsx"))
    if len(candidates) == 1:
        print(f"[note] '{base}' not found; using closest match: {candidates[0]}")
        return candidates[0]
    if len(candidates) > 1:
        sys.exit(f"Multiple PRERNA files in {dir_}, please specify exactly:\n  " +
                 "\n  ".join(candidates))
    sys.exit(f"File not found: {path}  (and no PRERNA_Full_Backup_*.xlsx in {dir_})")


def reclassify(path: str):
    path = _resolve_xlsx(path)
    app = create_app()
    with app.app_context():
        wb = load_workbook(filename=path, read_only=True, data_only=True)
        ws = None
        for name in wb.sheetnames:
            if "employee" in name.lower() or "master" in name.lower():
                ws = wb[name]; break
        if ws is None:
            ws = wb.active

        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h else "" for h in next(rows_iter)]
        idx = {h.lower(): i for i, h in enumerate(headers)}

        def col(row, key):
            i = idx.get(key.lower())
            return row[i] if i is not None and i < len(row) else None

        changed = unchanged = missing = 0
        change_log = []

        for row in rows_iter:
            if not row or all(v in (None, "") for v in row):
                continue
            emp_code = col(row, "Emp Code")
            if not emp_code:
                continue
            emp_code = str(emp_code).strip()
            prerna_role = (str(col(row, "Role") or "").strip()) or "EMPLOYEE"
            new_role = _role_for(prerna_role)
            full_name = str(col(row, "Full Name") or "").strip()

            u = User.query.filter_by(username=emp_code).first()
            if not u:
                missing += 1
                continue
            if u.role == new_role:
                unchanged += 1
                continue
            old_role = u.role
            u.role = new_role
            changed += 1
            change_log.append((emp_code, full_name, old_role, new_role, prerna_role))

        db.session.commit()
        wb.close()

        print()
        print(f"Re-classification complete:")
        print(f"  Changed   : {changed}")
        print(f"  Unchanged : {unchanged}")
        print(f"  Missing   : {missing}  (in PRERNA but no user row — re-run import_prerna.py)")
        print()
        if change_log:
            print("Role changes applied:")
            for emp, name, old, new, prerna in change_log:
                print(f"  {emp:12s} {name:30s} {old:10s} → {new:12s}  (PRERNA={prerna})")
            print()

        # Final role tally
        role_counts = Counter(u.role for u in User.query.all())
        print("Final role breakdown in DB:")
        for r, n in role_counts.most_common():
            print(f"  {n:4d} × {r}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        sys.exit(f"Usage: python {sys.argv[0]} /path/to/PRERNA_Full_Backup.xlsx")
    reclassify(sys.argv[1])
