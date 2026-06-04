"""Procam Attendance System — application factory."""
import os
from flask import Flask, redirect, url_for
from flask_login import LoginManager, current_user
from flask_migrate import Migrate

from config import Config
from models import (db, User, ROLE_ADMIN, ROLE_PROCAM_REP, ROLE_VENDOR_REP,
                    ROLE_WORKER, ROLE_GATE_GUARD)
from utils import indian_commas, amount_in_words

login_manager = LoginManager()
migrate = Migrate()


def create_app(config_class=Config):
    # instance_relative_config makes Flask create instance/ automatically
    app = Flask(__name__, instance_relative_config=True)
    app.config.from_object(config_class)

    # ensure instance dir exists
    try:
        os.makedirs(app.instance_path, exist_ok=True)
    except OSError:
        pass

    db.init_app(app)
    migrate.init_app(app, db)
    login_manager.init_app(app)
    login_manager.login_view = "auth.login"

    @login_manager.user_loader
    def load_user(uid):
        return User.query.get(int(uid))

    # Jinja helpers
    app.jinja_env.globals.update(
        APP_NAME=app.config["APP_NAME"],
        APP_TAGLINE=app.config["APP_TAGLINE"],
        BRAND_RED=app.config["BRAND_RED"],
        ROLE_ADMIN=ROLE_ADMIN,
        ROLE_PROCAM_REP=ROLE_PROCAM_REP,
        ROLE_VENDOR_REP=ROLE_VENDOR_REP,
        ROLE_WORKER=ROLE_WORKER,
    )
    app.jinja_env.filters["incomma"] = indian_commas
    app.jinja_env.filters["inwords"] = amount_in_words

    # Blueprints
    from routes.auth import bp as auth_bp
    from routes.admin import bp as admin_bp
    from routes.vendor import bp as vendor_bp
    from routes.client import bp as client_bp
    from routes.attendance import bp as attendance_bp
    from routes.invoice import bp as invoice_bp
    from routes.api import bp as api_bp
    from routes.kiosk import bp as kiosk_bp
    from routes.payroll import bp as payroll_bp

    app.register_blueprint(auth_bp)
    app.register_blueprint(admin_bp, url_prefix="/admin")
    app.register_blueprint(vendor_bp, url_prefix="/vendor")
    app.register_blueprint(client_bp, url_prefix="/client")
    app.register_blueprint(attendance_bp, url_prefix="/attendance")
    app.register_blueprint(invoice_bp, url_prefix="/invoice")
    app.register_blueprint(api_bp, url_prefix="/api/v1")
    app.register_blueprint(kiosk_bp)
    app.register_blueprint(payroll_bp)  # /payroll, /payroll/worker/<id>, /payroll/export.xlsx

    @app.route("/")
    def index():
        if not current_user.is_authenticated:
            return redirect(url_for("auth.login"))
        role = current_user.role
        # First-login force-chain: must change password first…
        if getattr(current_user, "must_change_password", False):
            return redirect(url_for("auth.change_password"))
        # …then EVERY employee linked to a Worker record must enrol their
        # face (Admin, ProcamRep, Worker — gate kiosk recognises everyone).
        # Gate guards don't punch, so they're exempt.
        if (role != ROLE_GATE_GUARD
                and getattr(current_user, "worker_id", None)):
            w = current_user.worker
            if w and not w.face_template:
                return redirect(url_for("attendance.enroll_face"))
        # Once enrolled (or if exempt), route by role to the right home page.
        if role == ROLE_ADMIN:
            return redirect(url_for("admin.dashboard"))
        if role == ROLE_PROCAM_REP:
            return redirect(url_for("client.dashboard"))
        if role == ROLE_VENDOR_REP:
            return redirect(url_for("vendor.dashboard"))
        if role == ROLE_GATE_GUARD:
            return redirect(url_for("kiosk.gate_screen"))
        return redirect(url_for("attendance.worker_home"))

    # auto-create tables on first boot + ensure an HR admin exists.
    # This is idempotent — safe to run every time the worker starts.
    with app.app_context():
        db.create_all()
        _patch_schema()
        _ensure_admin()

    # One-shot recovery endpoints — only live when SETUP_TOKEN env var is set
    _register_setup_routes(app)

    return app


def _ensure_admin():
    """Make sure at least one HR-admin user exists. Used on first deploy so the
       owner can sign in immediately. If admin already exists, no-op.

       Credentials come from env vars when set:
         ADMIN_USERNAME (default 'admin')
         ADMIN_PASSWORD (default 'admin123' — change on first login)
    """
    from models import User
    username = os.environ.get("ADMIN_USERNAME", "admin")
    password = os.environ.get("ADMIN_PASSWORD", "admin123")
    existing = User.query.filter_by(username=username).first()
    if existing:
        print(f"[bootstrap] admin user '{username}' already exists "
              f"(id={existing.id}, active={existing.is_active}, "
              f"hash_len={len(existing.password_hash or '')})")
    else:
        u = User(username=username, display_name="HR Admin",
                 role=ROLE_ADMIN, is_active=True, must_change_password=True)
        u.set_password(password)
        print(f"[bootstrap] hashing test: set→verify={u.check_password(password)}")
        db.session.add(u)
        try:
            db.session.commit()
            print(f"[bootstrap] admin user '{username}' created (must_change_password=True)")
        except Exception as e:
            db.session.rollback()
            print(f"[bootstrap] could not create admin: {type(e).__name__}: {e}")

    # Also ensure the gate kiosk guard exists. Same env-var override pattern.
    # We deliberately leave must_change_password=False because the kiosk is a
    # shared device manned by rotating staff — forcing each shift to set a
    # new password would be hostile UX.
    g_user = os.environ.get("GATE_USERNAME", "gate1")
    g_pass = os.environ.get("GATE_PASSWORD", "gate1")
    g_existing = User.query.filter_by(username=g_user).first()
    if g_existing:
        print(f"[bootstrap] gate user '{g_user}' already exists")
        return
    g = User(username=g_user, display_name="Gate 1 Security",
             role=ROLE_GATE_GUARD, is_active=True, must_change_password=False)
    g.set_password(g_pass)
    db.session.add(g)
    try:
        db.session.commit()
        print(f"[bootstrap] gate user '{g_user}' created (pw='{g_pass}')")
    except Exception as e:
        db.session.rollback()
        print(f"[bootstrap] could not create gate guard: {type(e).__name__}: {e}")


def _register_setup_routes(app):
    """One-shot recovery endpoints guarded by SETUP_TOKEN env var. Hit:
         /setup/health?token=<SETUP_TOKEN>
       to see whether admin exists, and
         /setup/reset-admin?token=<SETUP_TOKEN>&password=<new>
       to forcibly create-or-reset the admin user. Remove SETUP_TOKEN from
       Render's env vars once you're done — that disables these routes.
    """
    from flask import jsonify, request
    from models import User

    def _guard():
        # Read SETUP_TOKEN FRESH on every request (not captured at boot) so an
        # env-var change takes effect even if Render re-uses the worker process.
        # Strip whitespace from both sides — Render's UI sometimes preserves a
        # trailing newline / space when the value is pasted from a password
        # manager, which silently breaks exact-match comparison.
        expected = (os.environ.get("SETUP_TOKEN", "") or "").strip()
        token = (request.args.get("token") or "").strip()
        if not expected:
            return jsonify(error="setup routes disabled — set SETUP_TOKEN env var"), 403
        if token != expected:
            # Show lengths (not the values) so we can see whether there's a
            # hidden trailing character on either side.
            return jsonify(
                error="bad token",
                got_len=len(token),
                expected_len=len(expected),
                hint=("lengths differ — likely a hidden whitespace char in env var"
                      if len(token) != len(expected)
                      else "same length but value differs — env var content wrong"),
            ), 403
        return None

    @app.route("/setup/health")
    def setup_health():
        guard = _guard()
        if guard: return guard
        admin_username = os.environ.get("ADMIN_USERNAME", "admin")
        u = User.query.filter_by(username=admin_username).first()
        info = {
            "db_uri_kind": app.config["SQLALCHEMY_DATABASE_URI"].split("://", 1)[0],
            "total_users": User.query.count(),
            "looking_for": admin_username,
            "admin_exists": bool(u),
        }
        if u:
            info.update({
                "id": u.id, "is_active": u.is_active,
                "must_change_password": u.must_change_password,
                "role": u.role,
                "hash_length": len(u.password_hash or ""),
                "hash_prefix": (u.password_hash or "")[:7],
            })
        return jsonify(info)

    @app.route("/setup/check-user")
    def setup_check_user():
        """Diagnostic — does this username exist? Is it active? What role?
           Hit /setup/check-user?token=...&username=EMP1552018"""
        guard = _guard()
        if guard: return guard
        from sqlalchemy import func
        wanted = (request.args.get("username") or "").strip()
        if not wanted:
            return jsonify(error="pass ?username=<emp_code>"), 400
        # Case-insensitive lookup so we can spot capitalisation drift
        u = (User.query
             .filter(func.lower(User.username) == wanted.lower())
             .first())
        if not u:
            # Also report total count + any close-match suggestions
            sample = [r[0] for r in db.session.query(User.username)
                      .filter(User.username.ilike(f"%{wanted[:4]}%")).limit(5)]
            return jsonify(found=False, total_users=User.query.count(),
                           close_matches=sample), 404
        return jsonify(
            found=True, username_in_db=u.username,
            display_name=u.display_name, role=u.role,
            is_active=u.is_active, must_change_password=u.must_change_password,
            hash_length=len(u.password_hash or ""),
            hash_prefix=(u.password_hash or "")[:7],
            hint=("password is bcrypt-hashed; convention is password = username "
                  "with original capitalisation"),
        )

    @app.route("/setup/import-employees", methods=["GET", "POST"])
    def setup_import_employees():
        """Token-guarded PRERNA importer. GET shows an upload form, POST runs
           the import. Same logic as /admin/import-employees but doesn't require
           login — for first-time bootstrap when admin can't be reached.
           Hit /setup/import-employees?token=<SETUP_TOKEN>"""
        guard = _guard()
        if guard: return guard

        if request.method == "GET":
            return ("<!doctype html><html><body style='font-family:sans-serif; "
                    "max-width:600px; margin:60px auto; padding:20px;'>"
                    "<h2>Bootstrap PRERNA employees</h2>"
                    "<p>Upload PRERNA_Full_Backup_*.xlsx — imports/updates "
                    "every employee with the correct role.</p>"
                    f"<form method='post' enctype='multipart/form-data' "
                    f"action='/setup/import-employees?token={request.args.get('token','')}'>"
                    "<input type='file' name='file' accept='.xlsx,.xlsm' required "
                    "style='display:block; margin:20px 0; padding:10px; "
                    "border:2px dashed #BC1D2F; border-radius:6px; width:100%;'>"
                    "<button type='submit' style='padding:10px 20px; "
                    "background:#BC1D2F; color:#fff; border:none; border-radius:4px; "
                    "font-size:14px; cursor:pointer;'>Import everyone</button></form>"
                    "</body></html>")

        # POST — same logic as /admin/import-employees
        from collections import Counter
        from models import (Agency, Skill, Worker, ROLE_ADMIN, ROLE_PROCAM_REP,
                            ROLE_WORKER, APPROVER_NONE, CAT_SKILLED)
        import tempfile
        from datetime import date as _date
        from decimal import Decimal as _Dec

        ROLE_MAP = {"SUPER_ADMIN": ROLE_ADMIN, "HR_ADMIN": ROLE_ADMIN,
                    "MANAGER": ROLE_PROCAM_REP, "EMPLOYEE": ROLE_WORKER}

        f = request.files.get("file")
        if not f or not f.filename.lower().endswith((".xlsx", ".xlsm")):
            return jsonify(error="please POST a .xlsx file as 'file'"), 400

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tf:
            f.save(tf.name); tmp_path = tf.name

        try:
            from openpyxl import load_workbook
            wb = load_workbook(filename=tmp_path, read_only=True, data_only=True)
            ws = None
            for name in wb.sheetnames:
                if "employee" in name.lower() or "master" in name.lower():
                    ws = wb[name]; break
            if ws is None:
                ws = wb.active

            rows_iter = ws.iter_rows(values_only=True)
            headers = [str(h).strip() if h else "" for h in next(rows_iter)]
            ix = {h.lower(): i for i, h in enumerate(headers)}
            def col(row, key):
                i = ix.get(key.lower())
                return row[i] if i is not None and i < len(row) else None

            # Ensure PROCAM agency exists with mode=none
            ag = Agency.query.filter_by(code="PROCAM").first()
            if not ag:
                ag = Agency(code="PROCAM", name="Procam Logistics (In-house)",
                            contact_person="HR Admin", email="hr@procamlogistics.com",
                            address="Procam HO", default_gst_rate=_Dec("0.00"),
                            default_tds_rate=_Dec("0.00"),
                            onboarded_on=_date.today(), is_active=True,
                            approver_mode=APPROVER_NONE)
                db.session.add(ag); db.session.flush()
            elif ag.approver_mode != APPROVER_NONE:
                ag.approver_mode = APPROVER_NONE

            cw = uw = cu = uu = 0
            for row in rows_iter:
                if not row or all(v in (None, "") for v in row): continue
                emp_code = col(row, "Emp Code"); full_name = col(row, "Full Name")
                if not emp_code or not full_name: continue
                emp_code = str(emp_code).strip(); full_name = str(full_name).strip()

                vertical = (str(col(row, "Vertical") or "").strip())[:80] or None
                desig    = (str(col(row, "Designation") or "").strip())[:120] or "Employee"
                grade    = (str(col(row, "Grade") or "").strip())[:20] or None
                mgr_code = (str(col(row, "Mgr Code") or "").strip())[:32] or None
                prerna_role = (str(col(row, "Role") or "").strip()) or "EMPLOYEE"
                app_role = ROLE_MAP.get(prerna_role.strip().upper().replace(" ", "_"),
                                        ROLE_WORKER)

                sk = Skill.query.filter_by(name=desig).first()
                if not sk:
                    sk = Skill(name=desig, category=CAT_SKILLED)
                    db.session.add(sk); db.session.flush()

                w = Worker.query.filter_by(code=emp_code).first()
                if not w:
                    w = Worker(code=emp_code, full_name=full_name,
                               agency_id=ag.id, skill_id=sk.id, is_active=True,
                               onboarded_on=_date.today(), manager_code=mgr_code,
                               designation=desig, vertical=vertical, grade=grade)
                    db.session.add(w); db.session.flush()
                    cw += 1
                else:
                    w.full_name = full_name; w.agency_id = ag.id; w.skill_id = sk.id
                    w.is_active = True; w.manager_code = mgr_code
                    w.designation = desig; w.vertical = vertical; w.grade = grade
                    uw += 1

                u = User.query.filter_by(username=emp_code).first()
                if not u:
                    u = User(username=emp_code, display_name=full_name,
                             role=app_role, is_active=True,
                             must_change_password=True, worker_id=w.id)
                    u.set_password(emp_code)
                    db.session.add(u); cu += 1
                else:
                    u.display_name = full_name; u.worker_id = w.id
                    u.role = app_role; u.is_active = True
                    uu += 1
                if (cw + uw) % 50 == 0: db.session.commit()

            db.session.commit(); wb.close()
            roles = Counter(u.role for u in User.query.all())
            return jsonify(ok=True, workers_created=cw, workers_updated=uw,
                           users_created=cu, users_updated=uu,
                           roles_now=dict(roles),
                           note=("Every employee logs in with username = "
                                 "password = their emp_code. Forced to change "
                                 "password + enrol face on first login."))
        except Exception as e:
            db.session.rollback()
            return jsonify(ok=False, error=f"{type(e).__name__}: {e}"), 500
        finally:
            try: os.unlink(tmp_path)
            except OSError: pass

    @app.route("/setup/list-users")
    def setup_list_users():
        """First 200 users in the DB — for debugging which roles exist on Render.
           Hit /setup/list-users?token=..."""
        guard = _guard()
        if guard: return guard
        rows = User.query.order_by(User.username).limit(200).all()
        from collections import Counter
        roles = Counter(u.role for u in User.query.all())
        return jsonify(
            total=User.query.count(),
            roles=dict(roles),
            sample=[{"username": u.username, "role": u.role,
                     "active": u.is_active,
                     "must_change_password": u.must_change_password}
                    for u in rows],
        )

    @app.route("/setup/reset-all-workers")
    def setup_reset_all_workers():
        """Reset every Worker user's password back to their username
           (i.e. emp_code) and force a new-password reset on next login.
           Optional ?clear_faces=1 also wipes their face templates so they
           must re-enrol from scratch.
        """
        from models import ROLE_WORKER as RW, FaceTemplate
        guard = _guard()
        if guard: return guard
        clear_faces = request.args.get("clear_faces") in ("1", "true", "yes")
        workers = User.query.filter_by(role=RW).all()
        reset_n = 0
        for u in workers:
            u.set_password(u.username)       # pw == emp_code (PRERNA convention)
            u.must_change_password = True
            u.is_active = True
            reset_n += 1
        faces_wiped = 0
        if clear_faces:
            faces_wiped = FaceTemplate.query.delete()
        db.session.commit()
        return jsonify(
            ok=True, role_reset="Worker",
            users_reset=reset_n,
            face_templates_wiped=faces_wiped,
            note=("Every worker: username=pw=emp_code; will be forced to change "
                  "password on next login. " +
                  ("Faces wiped — they must re-enrol." if clear_faces else
                   "Existing face enrolments preserved."))
        )

    @app.route("/setup/reset-admin")
    def setup_reset_admin():
        guard = _guard()
        if guard: return guard
        from models import ROLE_ADMIN as RA
        username = os.environ.get("ADMIN_USERNAME", "admin")
        new_pw = request.args.get("password") or "admin123"
        u = User.query.filter_by(username=username).first()
        created = False
        if not u:
            u = User(username=username, display_name="HR Admin",
                     role=RA, is_active=True, must_change_password=True)
            db.session.add(u); created = True
        u.set_password(new_pw)
        u.is_active = True
        u.must_change_password = True
        db.session.commit()
        # Verify the hash works
        ok = u.check_password(new_pw)
        return jsonify(
            created=created, username=username,
            password_set=True, verify_after_set=ok,
            hash_prefix=u.password_hash[:7],
            message=("CREATED" if created else "RESET") + f" — sign in as '{username}' with the password you supplied.",
        )


def _patch_schema():
    """Lightweight at-boot schema patcher — adds new columns to existing tables
       so live DBs survive model changes without a wipe. No-op if everything is
       in sync. Mirrors the HRMS pattern.
    """
    from sqlalchemy import inspect, text
    try:
        inspector = inspect(db.engine)
    except Exception:
        return
    # Use TIMESTAMP on Postgres, DATETIME on SQLite — both work everywhere.
    dialect = db.engine.dialect.name
    DT = "TIMESTAMP" if dialect == "postgresql" else "DATETIME"
    expected_columns = [
        # (table_name, column_name, DDL fragment — must be valid on both dialects)
        ("skills", "category", "VARCHAR(20) DEFAULT 'Unskilled' NOT NULL"),
        ("face_templates", "pose_count", "INTEGER DEFAULT 1 NOT NULL"),
        ("attendance", "punch_in_at",  DT),
        ("attendance", "punch_out_at", DT),
        ("agencies", "approver_mode", "VARCHAR(16) DEFAULT 'hr_only' NOT NULL"),
        # PRERNA — manager linkage + designation/vertical/grade per worker
        ("workers", "manager_code", "VARCHAR(32)"),
        ("workers", "designation",  "VARCHAR(120)"),
        ("workers", "vertical",     "VARCHAR(80)"),
        ("workers", "grade",        "VARCHAR(20)"),
        # Face audit — snapshot photo of the Centre pose
        ("face_templates", "snapshot_b64", "TEXT"),
    ]
    for table, column, coldef in expected_columns:
        try:
            if not inspector.has_table(table):
                continue
            existing = {c["name"] for c in inspector.get_columns(table)}
            if column in existing:
                continue
            with db.engine.connect() as conn:
                conn.execute(text(f"ALTER TABLE {table} ADD COLUMN {column} {coldef}"))
                conn.commit()
            print(f"[schema] added {table}.{column}")
        except Exception as e:
            print(f"[schema] skipped {table}.{column}: {e}")


app = create_app()


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5055, debug=True)
