"""Import PROCAM in-house employees from PRERNA_Full_Backup.xlsx.

For each row in the Employee Master sheet:
  - Find/create the in-house Agency "Procam Logistics (In-house)" (code PROCAM)
  - Find/create a Skill from the Designation (default category Skilled)
  - Create a Worker (code=emp_code, full_name, agency=PROCAM, skill, is_active)
  - Create a User (username=emp_code, password=emp_code, role=Worker,
                  worker_id=<new>, must_change_password=True)

Idempotent: skips Workers/Users that already exist by emp_code.

Usage:
    python import_prerna.py "/path/to/PRERNA_Full_Backup.xlsx"
"""
import sys, os
from datetime import date
from decimal import Decimal

from openpyxl import load_workbook

from app import create_app
from models import (db, User, Agency, Skill, Worker,
                    ROLE_WORKER, CAT_SKILLED)

PROCAM_AGENCY_CODE = "PROCAM"
PROCAM_AGENCY_NAME = "Procam Logistics (In-house)"


def _ensure_agency():
    a = Agency.query.filter_by(code=PROCAM_AGENCY_CODE).first()
    if a:
        return a
    a = Agency(
        code=PROCAM_AGENCY_CODE, name=PROCAM_AGENCY_NAME,
        contact_person="Nilesh Sinha", email="hr@procamlogistics.com",
        address="731, Westend Mall, District Centre, Janakpuri, New Delhi-110058.",
        gstin=None, pan=None,
        default_gst_rate=Decimal("0.00"), default_tds_rate=Decimal("0.00"),
        onboarded_on=date.today(), is_active=True,
    )
    db.session.add(a); db.session.flush()
    print(f"  ✓ created in-house agency '{PROCAM_AGENCY_NAME}' (id={a.id})")
    return a


def _ensure_skill(name: str, category: str = CAT_SKILLED) -> Skill:
    name = (name or "Employee").strip()[:80] or "Employee"
    sk = Skill.query.filter_by(name=name).first()
    if sk:
        return sk
    sk = Skill(name=name, category=category)
    db.session.add(sk); db.session.flush()
    return sk


def import_prerna(path: str):
    if not os.path.isfile(path):
        sys.exit(f"File not found: {path}")
    app = create_app()
    with app.app_context():
        db.create_all()
        agency = _ensure_agency()

        wb = load_workbook(filename=path, read_only=True, data_only=True)
        # Pick the Employee Master sheet (case-insensitive match)
        ws = None
        for name in wb.sheetnames:
            if "employee" in name.lower() or "master" in name.lower():
                ws = wb[name]; break
        if ws is None:
            ws = wb.active
        print(f"Reading sheet: {ws.title!r}")

        # Map headers → column index
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h else "" for h in next(rows_iter)]
        idx = {h.lower(): i for i, h in enumerate(headers)}

        def col(row, key):
            i = idx.get(key.lower())
            return row[i] if i is not None and i < len(row) else None

        created_w = updated_w = created_u = updated_u = skipped = 0
        for row in rows_iter:
            if not row or all(v in (None, "") for v in row):
                continue
            emp_code = col(row, "Emp Code")
            full_name = col(row, "Full Name")
            if not emp_code or not full_name:
                continue
            emp_code = str(emp_code).strip()
            full_name = str(full_name).strip()

            vertical = (str(col(row, "Vertical") or "").strip())[:80] or None
            designation = (str(col(row, "Designation") or "").strip())[:80] or "Employee"
            grade = (str(col(row, "Grade") or "").strip())[:20] or None

            sk = _ensure_skill(designation, category=CAT_SKILLED)

            # Upsert Worker
            w = Worker.query.filter_by(code=emp_code).first()
            if not w:
                w = Worker(
                    code=emp_code, full_name=full_name,
                    agency_id=agency.id, skill_id=sk.id,
                    is_active=True, onboarded_on=date.today(),
                )
                db.session.add(w); db.session.flush()
                created_w += 1
            else:
                w.full_name = full_name
                w.agency_id = agency.id
                w.skill_id = sk.id
                w.is_active = True
                updated_w += 1

            # Upsert User
            u = User.query.filter_by(username=emp_code).first()
            if not u:
                u = User(
                    username=emp_code, display_name=full_name,
                    role=ROLE_WORKER, is_active=True,
                    must_change_password=True,   # forced first-login reset
                    worker_id=w.id,
                )
                u.set_password(emp_code)         # pw == username convention
                db.session.add(u)
                created_u += 1
            else:
                # If user exists but isn't linked to this worker, fix the link
                if u.worker_id != w.id:
                    u.worker_id = w.id
                u.display_name = full_name
                u.is_active = True
                updated_u += 1

            if (created_w + updated_w) % 50 == 0:
                db.session.commit()

        db.session.commit()
        wb.close()

        print()
        print(f"Workers : {created_w} created · {updated_w} updated")
        print(f"Users   : {created_u} created · {updated_u} updated")
        print()
        print("Login convention: username = employee code, password = employee code.")
        print("Every user must change password on first login, then enrol their face.")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        sys.exit(f"Usage: python {sys.argv[0]} /path/to/PRERNA_Full_Backup.xlsx")
    import_prerna(sys.argv[1])
